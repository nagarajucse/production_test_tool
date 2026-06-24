"""
routes.py — Flask Blueprint defining the HTTP DMS Server API endpoints.

Endpoints:
  POST /          — Accepts a JSON test result payload, validates it, persists to PostgreSQL.
  GET  /data      — Paginated query of all test results (for dashboard).
  GET  /stats     — Aggregate statistics for dashboard header cards.
  GET  /health    — Health check; returns server status and DB connectivity.
"""

import logging
import os
import base64
import io
import threading
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from datetime import datetime, timezone
from functools import wraps

from flask import Blueprint, Response, jsonify, request, session, redirect, url_for, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import select

from database import get_db
from models import SensorTestResult, User
from schemas import SensorTestResultSchema

logger = logging.getLogger("dms.routes")

api_bp = Blueprint("api", __name__)

excel_lock = threading.Lock()
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Datasheet.xlsx")

def append_to_excel(record: SensorTestResult):
    with excel_lock:
        if not os.path.exists(EXCEL_PATH):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Results"
            headers = ["Date/Time", "Sensor SN", "MAC Address", "Model", "AFIQ Score", "NFIQ Score", "Minutiae", "Work Order", "Tester", "Image Format", "Fingerprint Photo"]
            ws.append(headers)
        else:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            # Ensure headers exist
            if ws.max_row <= 1 and ws.cell(row=1, column=1).value is None:
                headers = ["Date/Time", "Sensor SN", "MAC Address", "Model", "AFIQ Score", "NFIQ Score", "Minutiae", "Work Order", "Tester", "Image Format", "Fingerprint Photo"]
                ws.append(headers)
        
        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            record.sensor_sn or "",
            record.sensor_mac or "",
            record.model or "",
            record.quality_score_afiq if record.quality_score_afiq is not None else "",
            record.nfiq_score if record.nfiq_score is not None else "",
            record.minutiae_count if record.minutiae_count is not None else "",
            record.work_order or "",
            record.tester_id or "",
            record.image_format or ""
        ]
        ws.append(row)
        
        # Add image
        if record.fingerprint_image:
            try:
                row_idx = ws.max_row
                img_data = io.BytesIO(record.fingerprint_image)
                pil_img = PILImage.open(img_data)
                orig_w, orig_h = pil_img.size
                ratio = 100.0 / orig_h
                new_w = int(orig_w * ratio)
                
                # Re-create BytesIO for openpyxl because PIL consumes it
                img_data.seek(0)
                xl_img = OpenpyxlImage(img_data)
                xl_img.width = new_w
                xl_img.height = 100
                
                col_letter = "K" # 11th column is Photo
                cell_ref = f"{col_letter}{row_idx}"
                ws.add_image(xl_img, cell_ref)
                
                ws.row_dimensions[row_idx].height = 80
                current_w = ws.column_dimensions[col_letter].width
                if current_w is None or current_w < (new_w / 7.0):
                    ws.column_dimensions[col_letter].width = (new_w / 7.0) + 2
            except Exception as e:
                logger.error("Error adding image to excel: %s", e)
                
        wb.save(EXCEL_PATH)


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            # For JSON API requests, return 401 Unauthorized
            if request.path in ["/data", "/stats"]:
                return jsonify({"status": "error", "message": "Unauthorized. Please log in."}), 401
            # For dashboard/page access, redirect to /login
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated_function


@api_bp.route("/favicon.ico", methods=["GET"])
@api_bp.route("/favicon.png", methods=["GET"])
def serve_favicon():
    favicon_path = os.path.join(os.path.dirname(__file__), "static", "favicon.png")
    return send_file(favicon_path, mimetype="image/png")


@api_bp.route("/login", methods=["GET"])
def serve_login():
    login_path = os.path.join(os.path.dirname(__file__), "login.html")
    return send_file(login_path, mimetype="text/html")


@api_bp.route("/register", methods=["POST"])
def register_user() -> tuple[Response, int]:
    payload = request.get_json(silent=True) or {}
    username = payload.get("username", "").strip()
    password = payload.get("password", "").strip()

    if not username or not password:
        return jsonify({"status": "error", "message": "Username and password are required."}), 400

    if len(username) < 3 or len(username) > 50:
        return jsonify({"status": "error", "message": "Username must be between 3 and 50 characters."}), 400

    if len(password) < 6:
        return jsonify({"status": "error", "message": "Password must be at least 6 characters long."}), 400

    try:
        with get_db() as db:
            # Check if user already exists
            existing_user = db.scalar(select(User).where(User.username == username))
            if existing_user:
                return jsonify({"status": "error", "message": "Username is already taken."}), 400

            # Hash the password and save
            pwd_hash = generate_password_hash(password)
            new_user = User(username=username, password_hash=pwd_hash)
            db.add(new_user)
            db.flush()
            user_id = str(new_user.id)

        logger.info("New user registered successfully: %s", username)
        return jsonify({"status": "ok", "message": "Registration successful. Please log in."}), 201
    except Exception as exc:
        logger.exception("User registration failed: %s", exc)
        return jsonify({"status": "error", "message": "An internal server error occurred."}), 500


@api_bp.route("/login", methods=["POST"])
def login_user() -> tuple[Response, int]:
    payload = request.get_json(silent=True) or {}
    username = payload.get("username", "").strip()
    password = payload.get("password", "").strip()

    if not username or not password:
        return jsonify({"status": "error", "message": "Username and password are required."}), 400

    try:
        with get_db() as db:
            user = db.scalar(select(User).where(User.username == username))
            if not user or not check_password_hash(user.password_hash, password):
                return jsonify({"status": "error", "message": "Invalid username or password."}), 401

            # Logged in successfully: set session variables
            session.clear()
            session.permanent = True
            session["user_id"] = str(user.id)
            session["username"] = user.username

        logger.info("User logged in successfully: %s", username)
        return jsonify({"status": "ok", "message": "Login successful."}), 200
    except Exception as exc:
        logger.exception("User login failed: %s", exc)
        return jsonify({"status": "error", "message": "An internal server error occurred."}), 500


@api_bp.route("/logout", methods=["POST"])
def logout_user() -> tuple[Response, int]:
    username = session.get("username", "Unknown")
    session.clear()
    logger.info("User logged out: %s", username)
    return jsonify({"status": "ok", "message": "Logged out successfully."}), 200


@api_bp.route("/dashboard", methods=["GET"])
@login_required
def serve_dashboard():
    dashboard_path = os.path.join(os.path.dirname(__file__), "dashboard.html")
    return send_file(dashboard_path, mimetype="text/html")

# ---------------------------------------------------------------------------
# POST / — Ingest a sensor test result
# ---------------------------------------------------------------------------

@api_bp.route("/", methods=["POST"])
def ingest_test_result() -> tuple[Response, int]:
    client_ip: str = request.remote_addr or "unknown"
    logger.info("POST / received from %s", client_ip)

    payload = request.get_json(silent=True)
    if payload is None:
        logger.warning("POST / rejected — missing or non-JSON body from %s", client_ip)
        return jsonify({
            "status": "error",
            "message": "Request body must be valid JSON with Content-Type: application/json.",
        }), 400

    try:
        validated = SensorTestResultSchema(**payload)
    except Exception as exc:
        logger.warning("POST / validation failed from %s: %s", client_ip, exc)
        return jsonify({
            "status": "error",
            "message": f"Payload validation failed: {exc}",
        }), 400

    try:
        # Extract and decode image data
        image_b64 = None
        for key in ["image", "fingerprint_image", "fingerprint", "image_data", "base64_image"]:
            if key in payload and isinstance(payload[key], str):
                image_b64 = payload[key]
                break

        image_name = payload.get("image_name")
        image_format = payload.get("image_format")

        fingerprint_image_bytes = None
        if image_b64:
            if image_b64.startswith("data:image"):
                try:
                    prefix, data_part = image_b64.split(",", 1)
                    if not image_format:
                        if "/" in prefix:
                            parts = prefix.split(";")[0].split("/")
                            if len(parts) > 1:
                                image_format = parts[1]
                    image_b64 = data_part
                except Exception:
                    pass

            try:
                fingerprint_image_bytes = base64.b64decode(image_b64)
            except Exception as e:
                logger.warning("Failed to decode base64 image: %s", e)

        if fingerprint_image_bytes and not image_format:
            image_format = "png"

        record = SensorTestResult(
            sensor_sn=validated.sensor_sn,
            model=validated.model,
            sensor_mac=validated.sensor_mac,
            quality_score_afiq=validated.quality_score_afiq,
            nfiq_score=validated.nfiq_score,
            minutiae_count=validated.minutiae_count,
            verification_score=validated.verification_score,
            part_number=validated.part_number,
            work_order=validated.work_order,
            tester_id=validated.tester_id,
            timestamp=validated.timestamp,
            received_at=datetime.now(timezone.utc),
            client_ip=client_ip,
            image_name=image_name,
            image_format=image_format,
            fingerprint_image=fingerprint_image_bytes,
        )
        # Capture values inside the session block to avoid DetachedInstanceError
        with get_db() as db:
            logger.info("===== DEBUG BEFORE INSERT =====")
            logger.info("image_name=%s", image_name)
            logger.info("image_format=%s", image_format)

            if fingerprint_image_bytes:
                logger.info(
                    "fingerprint_image_bytes length=%s",
                    len(fingerprint_image_bytes)
                )
            else:
                logger.error("fingerprint_image_bytes is None")
            
            db.add(record)
            db.flush()
            record_id = str(record.id)
            record_sn = record.sensor_sn
            record_wo = record.work_order

        logger.info("POST / stored record id=%s sensor_sn=%s work_order=%s", record_id, record_sn, record_wo)
        try:
            append_to_excel(record)
        except Exception as e:
            logger.error("Failed to append to excel: %s", e)
        return jsonify({
            "status": "ok",
            "message": "Test result stored.",
            "id": record_id,
        }), 200

    except Exception as exc:
        logger.exception("POST / database write failed from %s: %s", client_ip, exc)
        return jsonify({
            "status": "error",
            "message": "An internal server error occurred.",
        }), 500


# ---------------------------------------------------------------------------
# GET /image/<id> — Retrieve decoded fingerprint image for preview
# ---------------------------------------------------------------------------

@api_bp.route("/image/<record_id>", methods=["GET"])
@login_required
def get_image(record_id: str) -> Response:
    from database import get_db
    try:
        with get_db() as db:
            record = db.scalar(select(SensorTestResult).where(SensorTestResult.id == record_id))
            if not record:
                return jsonify({"status": "error", "message": "Record not found."}), 404
            
            image_bytes = record.fingerprint_image
            if not image_bytes:
                return jsonify({"status": "error", "message": "Image not found in record payload."}), 404
            
            logger.info("GET /image/%s", record_id)
            logger.info("image_size=%s", len(image_bytes))
            logger.info("image_format=%s", record.image_format)
            
            return send_file(
                io.BytesIO(image_bytes),
                mimetype=f"image/{record.image_format or 'png'}",
                as_attachment=False
            )
    except Exception as exc:
        logger.exception("GET /image/%s failed: %s", record_id, exc)
        return jsonify({"status": "error", "message": "Failed to retrieve image."}), 500


# ---------------------------------------------------------------------------
# GET /download/<id> — Download decoded fingerprint image
# ---------------------------------------------------------------------------

@api_bp.route("/download/<record_id>", methods=["GET"])
@login_required
def download_image(record_id: str) -> Response:
    from database import get_db
    try:
        with get_db() as db:
            record = db.scalar(select(SensorTestResult).where(SensorTestResult.id == record_id))
            if not record:
                return jsonify({"status": "error", "message": "Record not found."}), 404
            
            image_bytes = record.fingerprint_image
            if not image_bytes:
                return jsonify({"status": "error", "message": "Image not found in record payload."}), 404
            
            filename = record.image_name or f"fingerprint_{record.sensor_sn}_{record_id[:8]}.{record.image_format or 'png'}"
            
            return send_file(
                io.BytesIO(image_bytes),
                mimetype=f"image/{record.image_format or 'png'}",
                as_attachment=True,
                download_name=filename
            )
    except Exception as exc:
        logger.exception("GET /download/%s failed: %s", record_id, exc)
        return jsonify({"status": "error", "message": "Failed to download image."}), 500


# ---------------------------------------------------------------------------
# GET /data — Paginated query of all test results (for dashboard)
# ---------------------------------------------------------------------------

@api_bp.route("/data", methods=["GET"])
@login_required
def get_data() -> tuple[Response, int]:
    from database import get_engine
    from sqlalchemy import text as sa_text

    try:
        page     = max(1, int(request.args.get("page", 1)))
        per_page = min(200, max(1, int(request.args.get("per_page", 50))))
        search   = request.args.get("search", "").strip()
    except ValueError:
        return jsonify({"status": "error", "message": "Invalid pagination parameters."}), 400

    offset = (page - 1) * per_page
    where  = "WHERE (sensor_sn LIKE :q OR work_order LIKE :q OR tester_id LIKE :q)" if search else ""
    params: dict = {"limit": per_page, "offset": offset}
    if search:
        params["q"] = f"%{search}%"

    sql_rows = sa_text(f"""
        SELECT id, sensor_sn, model, sensor_mac,
               quality_score_afiq, nfiq_score, minutiae_count, verification_score,
               part_number, work_order, tester_id,
               timestamp, received_at, client_ip,
               image_name, image_format
        FROM sensor_test_results
        {where}
        ORDER BY received_at DESC
        LIMIT :limit OFFSET :offset
    """)
    sql_count = sa_text(f"SELECT COUNT(*) FROM sensor_test_results {where}")

    try:
        engine = get_engine()
        with engine.connect() as conn:
            total = conn.execute(sql_count, params).scalar()
            rows  = conn.execute(sql_rows, params).mappings().all()

        logger.info("GET /data returned %s rows", len(rows))
        if rows:
            logger.info("record=%s", rows[0]["id"])
            logger.info("image_name=%s", rows[0]["image_name"])
            logger.info("image_format=%s", rows[0]["image_format"])

        return jsonify({
            "status":   "ok",
            "total":    total,
            "page":     page,
            "per_page": per_page,
            "rows": [
                {
                    "id":                 str(r["id"]),
                    "sensor_sn":          r["sensor_sn"],
                    "model":              r["model"],
                    "sensor_mac":         r["sensor_mac"],
                    "quality_score_afiq": r["quality_score_afiq"],
                    "nfiq_score":         r["nfiq_score"],
                    "minutiae_count":     r["minutiae_count"],
                    "verification_score": r["verification_score"],
                    "part_number":        r["part_number"],
                    "work_order":         r["work_order"],
                    "tester_id":          r["tester_id"],
                    "timestamp":          r["timestamp"].isoformat() if r["timestamp"] else None,
                    "received_at":        r["received_at"].isoformat() if r["received_at"] else None,
                    "client_ip":          r["client_ip"],
                    "image_name":         r["image_name"],
                    "image_format":       r["image_format"],
                    "has_image":          bool(r["image_name"] or r["image_format"]),
                    "image_url":          f"/image/{r['id']}"
                }
                for r in rows
            ],
        }), 200

    except Exception as exc:
        logger.exception("GET /data failed: %s", exc)
        return jsonify({"status": "error", "message": "Failed to fetch data."}), 500


# ---------------------------------------------------------------------------
# GET /stats — Aggregate stats for dashboard header cards
# ---------------------------------------------------------------------------

@api_bp.route("/stats", methods=["GET"])
@login_required
def get_stats() -> tuple[Response, int]:
    from database import get_engine
    from sqlalchemy import text as sa_text

    sql = sa_text("""
        SELECT
            COUNT(*)                                    AS total_records,
            COUNT(DISTINCT sensor_sn)                   AS unique_sensors,
            COUNT(DISTINCT work_order)                  AS unique_work_orders,
            ROUND(AVG(quality_score_afiq), 1)  AS avg_quality,
            ROUND(AVG(nfiq_score), 1)          AS avg_nfiq,
            MAX(received_at)                            AS last_received
        FROM sensor_test_results
    """)
    try:
        engine = get_engine()
        with engine.connect() as conn:
            row = conn.execute(sql).mappings().one()
        return jsonify({
            "status":             "ok",
            "total_records":      row["total_records"],
            "unique_sensors":     row["unique_sensors"],
            "unique_work_orders": row["unique_work_orders"],
            "avg_quality":        float(row["avg_quality"] or 0),
            "avg_nfiq":           float(row["avg_nfiq"] or 0),
            "last_received":      row["last_received"].isoformat() if row["last_received"] else None,
        }), 200
    except Exception as exc:
        logger.exception("GET /stats failed: %s", exc)
        return jsonify({"status": "error", "message": "Failed to fetch stats."}), 500


# ---------------------------------------------------------------------------
# GET /health — Liveness + DB connectivity check
# ---------------------------------------------------------------------------

@api_bp.route("/health", methods=["GET"])
def health_check() -> tuple[Response, int]:
    from database import get_engine
    from sqlalchemy import text

    now = datetime.now(timezone.utc).isoformat()
    try:
        engine = get_engine()
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        logger.debug("GET /health — database OK")
        return jsonify({"status": "ok", "database": "connected", "timestamp": now}), 200
    except Exception as exc:
        logger.error("GET /health — database unreachable: %s", exc)
        return jsonify({"status": "error", "database": "unreachable", "timestamp": now}), 503

@api_bp.route("/download_datasheet", methods=["GET"])
@login_required
def download_datasheet():
    if not os.path.exists(EXCEL_PATH):
        return jsonify({"status": "error", "message": "Datasheet does not exist yet."}), 404
    return send_file(EXCEL_PATH, as_attachment=True, download_name="Datasheet.xlsx")